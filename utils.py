
import config
import os
import json
import traceback
from log import log
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment,Border, Side


#CRIA ARQUIVO JSON COM AS INFOS CAPTURADAS DOS ESTABELECIMENTOS
def response_file_json(data):
    try:
        log("ADICIONAR INFORMACOES NO ARQUIVO JSON")

        # GARANTE QUE SEJA UMA LISTA
        if isinstance(data, dict):
            data = [data]
        elif not isinstance(data, list):
            data = []

        # CRIA OU LIMPA, CASO JA TENHA, O ARQUIVO JSON
        with open(config.JSON_OUTPUT, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
            
        
        return "Arquivo JSON atualizado com sucesso!"

    except Exception as e:
        log("ERRO AO GERAR ARQUIVO JSON","ERROR")
        log(traceback.format_exc(),"DEBUG")
        return None
    
    

#PEGA O ARQUVO JSON GERADO E TRANSFORMAR EM UM ARQUIVO EXCEL
def excel_file():
    try:
        log("COVERTENDO ARQUIVO JSON EM UM ARQUIVO EXCEL")


        excel_output = config.EXCEL_OUTPUT
        json_output = config.JSON_OUTPUT
        
        #PEGANDO CONTEUDO DO JSON E CONVERTENDO PRO EXCEL
        file_json = pd.read_json(json_output, dtype=False)
        format_infos(file_json)
        file_json = file_json.fillna("N/A")
        file_json.to_excel(excel_output, index=False)
        
        #AJUSTANDO LARGURA DAS COLUNAS DO EXCEL
        wb = load_workbook(excel_output)
        ws = wb.active
        
        #ESTILIZANDO HEADER DO EXCEL
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        
        # APLICANDO ESTILIZACAO
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
        
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
        #PERCORRENDO PRA DEIXAR A LARGURA DA COLUNA DINAMICA
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        
        
            ws.column_dimensions[col_letter].width = max_length + 2
            
        wb.save(excel_output)
        return "Arquivo EXCEL Gerado com sucesso!"

    except Exception as e:
        log("ERRO AO GERAR ARQUIVO EXCEL","ERROR")
        log(traceback.format_exc(),"DEBUG")
        return None
    
def format_infos(file_json):
    file_json = pd.read_json(config.JSON_OUTPUT, encoding="utf-8")

    # Substituir textos de erro por vazio
    file_json["Nota do estabelecimento"] = file_json["Nota do estabelecimento"].replace(
        ["Avaliação não encontrada", None], ""
    )

    file_json["Quantidade de avaliações"] = file_json["Quantidade de avaliações"].replace(
        ["Quantidade de avaliação não encontrada", None], ""
    )

    # Converter nota 4,6 → 4.6
    file_json["Nota do estabelecimento"] = (
        file_json["Nota do estabelecimento"]
        .astype(str)
        .str.replace(",", ".", regex=False)
    )

    # Transformar em número (quem não der vira NaN)
    file_json["Nota do estabelecimento"] = pd.to_numeric(
        file_json["Nota do estabelecimento"], errors="coerce"
    )

    file_json["Quantidade de avaliações"] = pd.to_numeric(
        file_json["Quantidade de avaliações"], errors="coerce"
    )